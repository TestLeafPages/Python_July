# Step01:- install openpyxl:- pip install openpyxl

import openpyxl

# Load the workbook(excel)
wb = openpyxl.load_workbook(r"DataSheet.xlsx")

# returns the sheet names
print(wb.sheetnames)

# returns active sheet name
print(wb.active)

# Create the object for sheet
sheet = wb['datasheet00']

# for row count
print("total row count", sheet.max_row)


# for column count
print("total count count", sheet.max_column)

# Reading the value
val = sheet.cell(2, 2).value
# Way 01:
# for row in range(2, sheet.max_row + 1): # for row
#     for cell in range(1, sheet.max_column+1): # for cell
#         val = sheet.cell(row, cell).value
#         print(val)

# Way 02:
# print(sheet['B2'].value)
for row in sheet['A2' : 'C4']:
    for cell in row:
        print(cell.value)




